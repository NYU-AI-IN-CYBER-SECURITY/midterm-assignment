#!/usr/bin/env python3
"""
report.py -- writes the run report.  main.py feeds it one row at a time.

Output is an Excel workbook (results.xlsx) with three tabs:

    Summary   run settings, the binary and multiclass scores, per-class tables
    Results   one line per flow: truth vs prediction, correct?, seconds
    Prompts   one line per flow: the FULL prompt sent, and the raw model output

The Prompts tab is the one to read when your accuracy is bad: it shows exactly
what the model was asked and exactly what it said back, before any parsing.

Crash safety: rows are appended to a plain-text sidecar (.partial.jsonl) as
they happen and the workbook is assembled at the end, so killing a long run
never loses the rows that already finished.  The sidecar is deleted once the
workbook is written.

If openpyxl isn't installed, this falls back to two CSV files instead -- same
data, no tabs.
"""

import csv
import json
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:                       # degrade to CSV rather than crash
    HAVE_OPENPYXL = False

# Guard Rail: Excel refuses more than this many characters in one cell. 
CELL_LIMIT = 32_000
# guard rail: Excel also rejects most ASCII control characters inside a cell.
CONTROL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

ROW_HEADER = [
    "row", "truth_label", "pred_label", "label_correct",
    "truth_type", "pred_type", "type_correct",
    "parsed_ok", "elapsed_sec", "raw_output",
]
PROMPT_HEADER = ["row", "prompt_sent", "model_output",
                 "pred_label", "pred_type", "parsed_ok"]

HEAD_FILL = PatternFill("solid", fgColor="DDDDDD") if HAVE_OPENPYXL else None


def clean(value):
    """Make any value safe to put in a spreadsheet cell."""
    if not isinstance(value, str):
        return value
    text = CONTROL_CHARS.sub("", value)
    if len(text) > CELL_LIMIT:
        text = text[:CELL_LIMIT] + f"\n...[truncated, {len(value)} chars total]"
    return text


class Report:
    """Collects rows during the run, writes the workbook at the end."""

    def __init__(self, out_path: Path):
        self.rows = []            # Results tab
        self.prompts = []         # Prompts tab
        if HAVE_OPENPYXL:
            self.path = out_path.with_suffix(".xlsx")
        else:
            self.path = out_path.with_suffix(".csv")
            print("[report] openpyxl not installed -- writing CSV files instead "
                  "of a workbook.\n"
                  "         Run 'python install.py' again for the .xlsx report.",
                  flush=True)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.partial = self.path.with_suffix(".partial.jsonl")
        self._sidecar = self.partial.open("w", encoding="utf-8")

    # -- during the run ------------------------------------------------------

    def add(self, index, truth_label, pred_label, truth_type, pred_type,
            parsed_ok, elapsed, prompt, raw):
        self.rows.append([
            index, truth_label, pred_label, pred_label == truth_label,
            truth_type, pred_type, pred_type == truth_type,
            parsed_ok, round(elapsed, 3),
            raw.replace("\n", " ").replace("\r", " "),
        ])
        self.prompts.append([index, prompt, raw, pred_label, pred_type, parsed_ok])

        # Append-and-flush so a killed run still leaves everything so far.
        self._sidecar.write(json.dumps({
            "row": index, "truth_label": truth_label, "pred_label": pred_label,
            "truth_type": truth_type, "pred_type": pred_type,
            "parsed_ok": parsed_ok, "elapsed_sec": round(elapsed, 3),
            "prompt": prompt, "model_output": raw,
        }) + "\n")
        self._sidecar.flush()

    # -- at the end ----------------------------------------------------------

    def write(self, binary: dict, multiclass: dict, parse_failures: int,
              meta: dict) -> Path:
        self._sidecar.close()
        if HAVE_OPENPYXL:
            self._write_xlsx(binary, multiclass, parse_failures, meta)
        else:
            self._write_csvs(binary, multiclass, parse_failures, meta)
        self.partial.unlink(missing_ok=True)
        return self.path

    # -- xlsx ----------------------------------------------------------------

    def _write_xlsx(self, binary, multiclass, parse_failures, meta):
        wb = openpyxl.Workbook()

        self._summary_sheet(wb.active, binary, multiclass, parse_failures, meta)
        self._table_sheet(wb.create_sheet("Results"), ROW_HEADER, self.rows,
                          widths=[6, 12, 12, 13, 15, 15, 12, 10, 11, 60])
        self._table_sheet(wb.create_sheet("Prompts"), PROMPT_HEADER, self.prompts,
                          widths=[6, 90, 50, 12, 15, 10], wrap_cols=(2, 3))
        wb.save(self.path)

    def _summary_sheet(self, ws, binary, multiclass, parse_failures, meta):
        ws.title = "Summary"
        bold = Font(bold=True)

        def line(*values, header=False):
            ws.append(list(values))
            if header:
                for cell in ws[ws.max_row]:
                    cell.font = bold

        line("RUN", header=True)
        for key, value in meta.items():
            line(key, value)
        line("parse_failures", parse_failures)

        for title, m in (("BINARY  (normal vs attack)", binary),
                         ("MULTICLASS  (attack type)", multiclass)):
            line()
            line(title, header=True)
            line("scored_rows", m["n"])
            line("correct", m["correct"])
            line("accuracy", m["accuracy"])
            line("macro_precision", m["macro_precision"])
            line("macro_recall", m["macro_recall"])
            line("macro_f1", m["macro_f1"])
            line()
            line("class", "support", "correct", "precision", "recall", "f1",
                 header=True)
            for cls, s in sorted(m["per_class"].items()):
                line(cls, s["support"], s["correct"],
                     s["precision"], s["recall"], s["f1"])

        for column, width in zip("ABCDEF", (18, 10, 10, 11, 10, 10)):
            ws.column_dimensions[column].width = width
        ws.freeze_panes = "A2"

    def _table_sheet(self, ws, header, rows, widths, wrap_cols=()):
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEAD_FILL
        for row in rows:
            ws.append([clean(v) for v in row])
            # openpyxl treats a string starting with "=" as a formula, and a
            # student's prompt may well start with one.  Force text.
            for cell in ws[ws.max_row]:
                if isinstance(cell.value, str):
                    cell.data_type = "s"
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for i in wrap_cols:
            for cell in ws[get_column_letter(i)][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    # -- csv fallback --------------------------------------------------------

    def _write_csvs(self, binary, multiclass, parse_failures, meta):
        with self.path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(ROW_HEADER)
            w.writerows(self.rows)
            w.writerow([])
            w.writerow(["=== RUN ==="])
            for key, value in meta.items():
                w.writerow([key, value])
            w.writerow(["parse_failures", parse_failures])
            for title, m in (("=== BINARY  (normal vs attack) ===", binary),
                             ("=== MULTICLASS  (attack type) ===", multiclass)):
                w.writerow([])
                w.writerow([title])
                for key in ("n", "correct", "accuracy", "macro_precision",
                            "macro_recall", "macro_f1"):
                    w.writerow([key, m[key]])
                w.writerow([])
                w.writerow(["class", "support", "correct", "precision", "recall", "f1"])
                for cls, s in sorted(m["per_class"].items()):
                    w.writerow([cls, s["support"], s["correct"],
                                s["precision"], s["recall"], s["f1"]])

        prompts_path = self.path.with_name(self.path.stem + "_prompts.csv")
        with prompts_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(PROMPT_HEADER)
            w.writerows(self.prompts)
        print(f"[report] prompts written: {prompts_path}", flush=True)
